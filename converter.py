from __future__ import annotations

import csv
import json
import tempfile
import unicodedata
from pathlib import Path
from typing import Any, Iterable, List, Dict

import httpx
from openpyxl import Workbook, load_workbook
from bs4 import BeautifulSoup
import xml.etree.ElementTree as ET


# ---------------------------------------------------------------------------
# 1. CSV -> JSON
# ---------------------------------------------------------------------------


def convert_csv_to_json(
    csv_path: str | Path,
    *,
    delimiter: str = ",",
    encoding: str = "utf-8",
    has_header: bool = True,
    pretty: bool = False,
    max_rows: int = 100_000,
) -> tuple[Path, int, str | None, list[dict[str, Any]]]:
    """
    Convertit un fichier CSV en JSON.
    Retourne (json_path, rows_count, warning, data_list).
    """
    csv_path = Path(csv_path)

    try:
        f = csv_path.open("r", encoding=encoding, newline="")
    except FileNotFoundError as e:
        raise ValueError(f"Fichier CSV introuvable : {csv_path}") from e

    with f:
        reader = csv.reader(f, delimiter=delimiter)
        rows = list(reader)

    if not rows:
        raise ValueError("Le fichier CSV est vide.")

    warning: str | None = None
    data: list[dict[str, Any]] = []

    if has_header:
        header = rows[0]
        if not header:
            raise ValueError("La ligne d'en-tête du CSV est vide.")

        for idx, row in enumerate(rows[1:], start=2):
            if idx - 1 > max_rows:
                warning = f"Fichier tronqué à {max_rows} lignes pour des raisons de performance."
                break

            if len(row) != len(header):
                warning = (
                    "Certaines lignes n'ont pas le même nombre de colonnes que l'en-tête. "
                    "Les valeurs manquantes sont complétées par des chaînes vides."
                )
            row = (row + [""] * len(header))[: len(header)]
            obj = {header[i]: row[i] for i in range(len(header))}
            data.append(obj)
    else:
        max_len = max(len(r) for r in rows)
        header = [f"col_{i+1}" for i in range(max_len)]
        warning = (
            "CSV sans en-tête : des noms de colonnes génériques col_1, col_2, ... ont été utilisés."
        )
        for idx, row in enumerate(rows, start=1):
            if idx > max_rows:
                warning = f"Fichier tronqué à {max_rows} lignes pour des raisons de performance."
                break
            row = (row + [""] * max_len)[:max_len]
            obj = {header[i]: row[i] for i in range(max_len)}
            data.append(obj)

    rows_count = len(data)

    tmp = tempfile.NamedTemporaryFile(
        delete=False, suffix=".json", mode="w", encoding="utf-8"
    )
    if pretty:
        json.dump(data, tmp, ensure_ascii=False, indent=2)
    else:
        json.dump(data, tmp, ensure_ascii=False, separators=(",", ":"))
    tmp.close()

    return Path(tmp.name), rows_count, warning, data


# ---------------------------------------------------------------------------
# 2. JSON -> CSV
# ---------------------------------------------------------------------------


def convert_json_to_csv(
    json_text: str,
    *,
    delimiter: str = ",",
    has_header: bool = True,
    max_rows: int = 100_000,
    flatten_nested: bool = False,
) -> tuple[Path, int, str | None]:
    """
    Convertit une liste JSON en CSV.
    """
    try:
        data = json.loads(json_text)
    except json.JSONDecodeError as e:
        raise ValueError(f"JSON invalide : {e}") from e

    if not isinstance(data, list) or len(data) == 0:
        raise ValueError("Le JSON doit être une liste non vide (d'objets ou de tableaux).")

    warning: str | None = None
    rows_count = 0

    tmp = tempfile.NamedTemporaryFile(
        delete=False, suffix=".csv", mode="w", newline="", encoding="utf-8"
    )
    writer = csv.writer(tmp, delimiter=delimiter)

    first = data[0]

    if isinstance(first, dict):
        keys: list[str] = []
        seen = set()
        # Première passe : collecter toutes les clés
        for obj in data:
            if rows_count >= max_rows:
                warning = f"Fichier tronqué à {max_rows} lignes pour des raisons de performance."
                break
            if not isinstance(obj, dict):
                raise ValueError("Tous les éléments doivent être des objets JSON (dict).")
            if flatten_nested:
                flat = {}
                for k, v in obj.items():
                    if isinstance(v, (dict, list)):
                        flat[k] = json.dumps(v, ensure_ascii=False)
                    else:
                        flat[k] = v
                obj = flat
            for k in obj.keys():
                if k not in seen:
                    seen.add(k)
                    keys.append(k)

        if has_header:
            writer.writerow(keys)

        rows_count = 0
        for obj in data:
            if rows_count >= max_rows:
                break
            if not isinstance(obj, dict):
                continue
            if flatten_nested:
                flat = {}
                for k, v in obj.items():
                    if isinstance(v, (dict, list)):
                        flat[k] = json.dumps(v, ensure_ascii=False)
                    else:
                        flat[k] = v
                obj = flat
            row = [obj.get(k, "") for k in keys]
            writer.writerow(row)
            rows_count += 1

    elif isinstance(first, (list, tuple)):
        if has_header:
            warning = "has_header=true est ignoré pour une liste de tableaux."
        for row in data:
            if rows_count >= max_rows:
                warning = f"Fichier tronqué à {max_rows} lignes pour des raisons de performance."
                break
            if not isinstance(row, (list, tuple)):
                raise ValueError(
                    "Tous les éléments doivent être des tableaux si le premier est un tableau."
                )
            writer.writerow(list(row))
            rows_count += 1
    else:
        raise ValueError(
            "Le JSON doit être une liste d'objets ({...}) ou une liste de tableaux ([...])."
        )

    tmp.close()
    return Path(tmp.name), rows_count, warning


# ---------------------------------------------------------------------------
# 3. CSV -> Excel (.xlsx)
# ---------------------------------------------------------------------------


def csv_to_excel(
    csv_path: str | Path,
    *,
    delimiter: str = ",",
    encoding: str = "utf-8",
    has_header: bool = True,
    sheet_name: str = "Sheet1",
    max_rows: int = 100_000,
) -> tuple[Path, int, str | None]:
    """
    Convertit un CSV en fichier Excel (.xlsx).
    """
    csv_path = Path(csv_path)
    try:
        f = csv_path.open("r", encoding=encoding, newline="")
    except FileNotFoundError as e:
        raise ValueError(f"Fichier CSV introuvable : {csv_path}") from e

    warning: str | None = None
    with f:
        reader = csv.reader(f, delimiter=delimiter)
        rows = []
        for idx, row in enumerate(reader, start=1):
            if idx > max_rows:
                warning = f"Fichier tronqué à {max_rows} lignes pour des raisons de performance."
                break
            rows.append(row)

    if not rows:
        raise ValueError("Le fichier CSV est vide.")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name or "Sheet1"

    for row in rows:
        ws.append(row)

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.close()

    rows_count = len(rows) - (1 if has_header else 0)
    return Path(tmp.name), rows_count, warning


# ---------------------------------------------------------------------------
# 4. Excel (.xlsx) -> CSV
# ---------------------------------------------------------------------------


def excel_to_csv(
    xlsx_path: str | Path,
    *,
    delimiter: str = ",",
    sheet_name: str | None = None,
    has_header: bool = True,
    max_rows: int = 100_000,
) -> tuple[Path, int, str | None]:
    """
    Convertit un fichier Excel (.xlsx) en CSV.
    """
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise ValueError(f"Fichier Excel introuvable : {xlsx_path}")

    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"Impossible de lire le fichier Excel : {e}") from e

    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    tmp = tempfile.NamedTemporaryFile(
        delete=False, suffix=".csv", mode="w", newline="", encoding="utf-8"
    )
    writer = csv.writer(tmp, delimiter=delimiter)

    warning: str | None = None
    rows_count = 0
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if idx > max_rows:
            warning = f"Fichier tronqué à {max_rows} lignes pour des raisons de performance."
            break
        writer.writerow([("" if v is None else v) for v in row])
        rows_count += 1

    tmp.close()
    if has_header and rows_count > 0:
        rows_count -= 1

    return Path(tmp.name), rows_count, warning


# ---------------------------------------------------------------------------
# 5. JSON Formatter (pretty / compact / validate)
# ---------------------------------------------------------------------------


def json_formatter(
    json_text: str,
    *,
    mode: str = "pretty",  # "pretty" ou "compact"
    indent: int = 2,
    sort_keys: bool = False,
    ensure_ascii: bool = False,
    validate: bool = True,
    max_size_kb: int = 1024,
) -> dict[str, Any]:
    """
    Formate du JSON (pretty/compact) et optionnellement le valide.
    """
    size_kb = len(json_text.encode("utf-8")) / 1024
    if size_kb > max_size_kb:
        raise ValueError(
            f"JSON trop volumineux ({size_kb:.1f} KB > {max_size_kb} KB autorisés)."
        )

    valid = True
    error: str | None = None
    data: Any | None = None

    try:
        data = json.loads(json_text)
    except json.JSONDecodeError as e:
        if validate:
            valid = False
            error = str(e)
        else:
            raise ValueError(f"JSON invalide : {e}") from e

    formatted: str | None = None
    if data is not None:
        if mode == "compact":
            formatted = json.dumps(
                data,
                ensure_ascii=ensure_ascii,
                separators=(",", ":"),
                sort_keys=sort_keys,
            )
        else:
            formatted = json.dumps(
                data,
                ensure_ascii=ensure_ascii,
                indent=indent,
                sort_keys=sort_keys,
            )

    return {
        "valid": valid,
        "error": error,
        "formatted": formatted,
        "mode": mode,
    }


# ---------------------------------------------------------------------------
# 6. XML -> JSON
# ---------------------------------------------------------------------------


def _xml_element_to_dict(
    elem: ET.Element,
    *,
    strip_whitespace: bool,
    max_depth: int,
    max_nodes: int,
    state: dict[str, int],
    text_key: str = "#text",
) -> dict[str, Any]:
    if state["nodes"] >= max_nodes:
        raise ValueError(f"XML trop volumineux (> {max_nodes} noeuds).")
    state["nodes"] += 1

    if max_depth <= 0:
        return {elem.tag: None}

    result: dict[str, Any] = {}
    # attributs
    if elem.attrib:
        result["@attributes"] = dict(elem.attrib)

    # texte
    text = elem.text or ""
    if strip_whitespace:
        text = text.strip()
    if text:
        result[text_key] = text

    # enfants
    children = list(elem)
    if children:
        child_dict: dict[str, Any] = {}
        for child in children:
            child_repr = _xml_element_to_dict(
                child,
                strip_whitespace=strip_whitespace,
                max_depth=max_depth - 1,
                max_nodes=max_nodes,
                state=state,
                text_key=text_key,
            )
            tag = child.tag
            value = child_repr
            if tag in child_dict:
                if not isinstance(child_dict[tag], list):
                    child_dict[tag] = [child_dict[tag]]
                child_dict[tag].append(value)
            else:
                child_dict[tag] = value
        result["children"] = child_dict

    return result


def xml_to_json(
    xml_text: str,
    *,
    strip_whitespace: bool = True,
    max_depth: int = 10,
    max_nodes: int = 10_000,
    text_key: str = "#text",
) -> dict[str, Any]:
    """
    Convertit du XML en structure JSON (dict).
    """
    try:
        root = ET.fromstring(xml_text)
    except ET.ParseError as e:
        raise ValueError(f"XML invalide : {e}") from e

    state = {"nodes": 0}
    result = _xml_element_to_dict(
        root,
        strip_whitespace=strip_whitespace,
        max_depth=max_depth,
        max_nodes=max_nodes,
        state=state,
        text_key=text_key,
    )
    return {root.tag: result}


# ---------------------------------------------------------------------------
# 7. JSON -> XML
# ---------------------------------------------------------------------------


def _dict_to_xml_element(
    tag: str,
    data: Any,
    *,
    attr_prefix: str = "@",
    text_key: str = "#text",
) -> ET.Element:
    elem = ET.Element(tag)

    if isinstance(data, dict):
        for k, v in data.items():
            if k.startswith(attr_prefix):
                attr_name = k[len(attr_prefix) :]
                elem.set(attr_name, str(v))
            elif k == text_key:
                elem.text = str(v)
            else:
                if isinstance(v, list):
                    for item in v:
                        child = _dict_to_xml_element(k, item, attr_prefix=attr_prefix, text_key=text_key)
                        elem.append(child)
                else:
                    child = _dict_to_xml_element(k, v, attr_prefix=attr_prefix, text_key=text_key)
                    elem.append(child)
    elif isinstance(data, list):
        for item in data:
            child = _dict_to_xml_element("item", item, attr_prefix=attr_prefix, text_key=text_key)
            elem.append(child)
    else:
        elem.text = "" if data is None else str(data)

    return elem


def json_to_xml(
    json_text: str,
    *,
    root_tag: str = "root",
    attr_prefix: str = "@",
    text_key: str = "#text",
    pretty: bool = True,
) -> str:
    """
    Convertit un JSON (dict ou liste) en XML.
    """
    try:
        data = json.loads(json_text)
    except json.JSONDecodeError as e:
        raise ValueError(f"JSON invalide : {e}") from e

    root = _dict_to_xml_element(root_tag, data, attr_prefix=attr_prefix, text_key=text_key)

    xml_str = ET.tostring(root, encoding="unicode")
    if not pretty:
        return xml_str

    # simple pretty-print (indentation)
    try:
        import xml.dom.minidom as minidom

        dom = minidom.parseString(xml_str.encode("utf-8"))
        return dom.toprettyxml(indent="  ")
    except Exception:
        return xml_str


# ---------------------------------------------------------------------------
# 8. HTML Table -> JSON
# ---------------------------------------------------------------------------


def html_table_to_json(
    html_text: str,
    *,
    table_index: int = 0,
    has_header: bool = True,
    convert_numbers: bool = True,
    max_rows: int = 10_000,
) -> list[dict[str, Any]] | list[list[Any]]:
    """
    Convertit un tableau HTML en JSON.
    """
    soup = BeautifulSoup(html_text, "html.parser")
    tables = soup.find_all("table")
    if not tables:
        raise ValueError("Aucun tableau <table> trouvé dans le HTML.")

    if table_index < 0 or table_index >= len(tables):
        raise ValueError(f"Index de tableau invalide (0 à {len(tables)-1}).")

    table = tables[table_index]
    rows = table.find_all("tr")
    if not rows:
        raise ValueError("Le tableau HTML ne contient aucune ligne.")

    def maybe_num(s: str) -> Any:
        if not convert_numbers:
            return s
        s_strip = s.strip().replace(",", ".")
        try:
            return float(s_strip) if s_strip else s
        except ValueError:
            return s

    data: list[Any] = []

    if has_header:
        header_cells = rows[0].find_all(["th", "td"])
        header = [c.get_text(strip=True) for c in header_cells]
        for row in rows[1 : 1 + max_rows]:
            cells = row.find_all(["th", "td"])
            values = [maybe_num(c.get_text(strip=True)) for c in cells]
            values = (values + [""] * len(header))[: len(header)]
            obj = {header[i]: values[i] for i in range(len(header))}
            data.append(obj)
        return data

    for row in rows[:max_rows]:
        cells = row.find_all(["th", "td"])
        values = [maybe_num(c.get_text(strip=True)) for c in cells]
        data.append(values)

    return data


# ---------------------------------------------------------------------------
# 9. CSV URL Fetcher (CSV URL -> JSON)
# ---------------------------------------------------------------------------


def csv_url_to_json(
    url: str,
    *,
    delimiter: str = ",",
    encoding: str = "utf-8",
    has_header: bool = True,
    pretty: bool = False,
    max_rows: int = 100_000,
    timeout: float = 10.0,
) -> tuple[Path, int, str | None, list[dict[str, Any]]]:
    """
    Télécharge un CSV depuis une URL et le convertit en JSON.
    """
    try:
        resp = httpx.get(url, timeout=timeout)
    except Exception as e:
        raise ValueError(f"Erreur réseau lors de la requête : {e}") from e

    if resp.status_code >= 400:
        raise ValueError(f"Requête HTTP échouée ({resp.status_code}).")

    content = resp.content
    tmp_csv = tempfile.NamedTemporaryFile(delete=False, suffix=".csv")
    tmp_csv.write(content)
    tmp_csv.close()

    return convert_csv_to_json(
        tmp_csv.name,
        delimiter=delimiter,
        encoding=encoding,
        has_header=has_header,
        pretty=pretty,
        max_rows=max_rows,
    )


# ---------------------------------------------------------------------------
# 10. Text Cleaner
# ---------------------------------------------------------------------------


def clean_text(
    text: str,
    *,
    trim: bool = True,
    normalize_unicode: bool = True,
    remove_accents: bool = False,
    collapse_whitespace: bool = True,
    to_lower: bool = False,
    max_length: int = 10_000,
) -> dict[str, Any]:
    """
    Nettoie un texte (trim, unicode, accents, espaces...).
    """
    original_length = len(text)

    if max_length and original_length > max_length:
        text = text[:max_length]

    if normalize_unicode:
        text = unicodedata.normalize("NFC", text)

    if remove_accents:
        text = "".join(
            c
            for c in unicodedata.normalize("NFD", text)
            if unicodedata.category(c) != "Mn"
        )

    if trim:
        text = text.strip()

    if collapse_whitespace:
        import re

        text = re.sub(r"\s+", " ", text)

    if to_lower:
        text = text.lower()

    cleaned_length = len(text)
    return {
        "original_length": original_length,
        "cleaned_length": cleaned_length,
        "cleaned": text,
    }
